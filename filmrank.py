#Reads an Excel file containing a list of movies in column A and their corresponding year in column B
#Extracts the ratings for each film from a movie ratings website
#Presents these ratings on a graph

import plotly
import plotly.graph_objs as go
import os
from openpyxl import load_workbook
import sys
import re
import urllib
import urlparse
import bs4

#Change these lines accordingly
os.chdir('C:\Users\Finian\Desktop')
wb = load_workbook(filename = 'DVDs.xlsx')
ws = wb['Sheet1']

class MyOpener(urllib.FancyURLopener):
    version = 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.2.15) Gecko/20110303 Firefox/3.6.15'

myopener = MyOpener()


def findRating():
    sauce = myopener.open(p1).read()
    soup = bs4.BeautifulSoup(sauce, 'lxml')
    res = soup.find('span', class_='meter-value superPageFontColor').span.text
    return int(res)

ratings = []
films = []

for row in range(ws.max_row):
    film = str(ws['A'+str(row+1)].value)
    l = []
    for k in film.split(): l.append(re.sub(r"[^a-zA-Z0-9\-]+", '', k))
    outFilm = film
    film = ' '.join(l)
    #film has no special chars

    #With year
    p1 = "https://www.rottentomatoes.com/m/" + film.replace(' ', '_') + '_' + str(ws['B'+str(row+1)].value)
    try:
        rating = findRating()
    except AttributeError:
        try:
            # Film without year
            p1 = "https://www.rottentomatoes.com/m/" + film.replace(' ', '_')
            rating = findRating()
        except:
            #Film without 1st word
            film1 = film.split(' ')
            film1.pop(0)
            if len(film1) > 1:
                film1 = '_'.join(film1)
            else:
                film1 = ''.join(film1)
            p1 = "https://www.rottentomatoes.com/m/" + film1
            rating = findRating()
    except:
        #Film with year - 1
        p1 = "https://www.rottentomatoes.com/m/" + film.replace(' ', '_') + '_' + str(ws['B'+str(row+1)].value-1)
        rating = findRating()
    ratings.append(rating)
    films.append(outFilm)

plotly.offline.plot({
    "data": [go.Scatter(x=films, y=ratings, mode = 'markers')],
    "layout": go.Layout(title="DVD Ratings"),
})

